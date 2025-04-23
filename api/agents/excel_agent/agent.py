from typing import Any, Dict, List, Optional, Literal, AsyncIterable
from pydantic import BaseModel
import pandas as pd
import os
import uuid
from pathlib import Path
from datetime import datetime

from core.models import ExcelRequestMessage, QueryResponse, SQLResultMessage
from core.config import settings

class ExcelFormat(BaseModel):
    """Excel formatting options"""
    sheet_name: str = "Data"
    include_headers: bool = True
    auto_filter: bool = True
    freeze_panes: bool = True
    column_width_auto: bool = True
    include_timestamp: bool = True
    include_query: bool = True
    include_metadata: bool = True
    style_template: Literal["default", "professional", "minimal", "colorful"] = "default"
    
class ExcelAgent:
    """Agent for converting SQL results to Excel files"""
    
    SUPPORTED_CONTENT_TYPES = ["text", "text/plain", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]
    
    def __init__(self, output_dir: str = None):
        self.output_dir = output_dir or os.path.join(os.getcwd(), "outputs", "excel")
        os.makedirs(self.output_dir, exist_ok=True)
    
    def process_request(self, request: ExcelRequestMessage) -> Dict[str, Any]:
        """Process an Excel request and return the path to the generated Excel file"""
        # Extract data from request
        query = request.query
        sql_query = request.sql_query
        result_data = request.result
        format_options = request.format_options or {}
        
        # Create format settings
        excel_format = ExcelFormat(**format_options)
        
        # Generate Excel file
        excel_path = self._generate_excel(
            query=query,
            sql_query=sql_query,
            data=result_data,
            format_options=excel_format
        )
        
        # Return result
        return {
            "is_task_complete": True,
            "require_user_input": False,
            "content": {
                "file_path": excel_path,
                "file_name": os.path.basename(excel_path),
                "timestamp": datetime.now().isoformat(),
                "format": excel_format.model_dump()
            }
        }
    
    def _generate_excel(self, query: str, sql_query: str, data: List[Dict[str, Any]], 
                        format_options: ExcelFormat) -> str:
        """Generate an Excel file from SQL results"""
        # Convert data to DataFrame
        df = pd.DataFrame(data)
        if df.empty:
            # Handle empty result
            df = pd.DataFrame({"No results": ["No data returned from query"]})
        
        # Generate unique filename
        filename = f"query_result_{uuid.uuid4().hex[:8]}.xlsx"
        file_path = os.path.join(self.output_dir, filename)
        
        # Create Excel writer
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Write data to Excel
            df.to_excel(
                writer, 
                sheet_name=format_options.sheet_name,
                index=False,
                freeze_panes=(1, 0) if format_options.freeze_panes else None
            )
            
            # Apply formatting
            self._apply_formatting(writer, df, format_options)
            
            # Add metadata sheet if requested
            if format_options.include_metadata:
                self._add_metadata_sheet(writer, query, sql_query, data)
        
        return file_path
    
    def _apply_formatting(self, writer: pd.ExcelWriter, df: pd.DataFrame, format_options: ExcelFormat):
        """Apply formatting to the Excel worksheet"""
        worksheet = writer.sheets[format_options.sheet_name]
        
        # Auto-filter
        if format_options.auto_filter and not df.empty:
            worksheet.auto_filter.ref = worksheet.dimensions
        
        # Adjust column widths
        if format_options.column_width_auto and not df.empty:
            for i, column in enumerate(df.columns):
                # Calculate width based on column name and maximum data length
                max_length = max(
                    df[column].astype(str).map(len).max(),
                    len(str(column))
                )
                # Add some padding
                adjusted_width = max_length + 2
                # Set column width
                col_letter = worksheet.cell(row=1, column=i+1).column_letter
                worksheet.column_dimensions[col_letter].width = adjusted_width
        
        # Apply style template
        self._apply_style_template(worksheet, format_options.style_template)
    
    def _apply_style_template(self, worksheet, template_name: str):
        """Apply a predefined style template to the worksheet"""
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        if template_name == "professional":
            # Header formatting
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
        elif template_name == "colorful":
            # Colorful alternating rows
            light_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            
            # Format header
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Format data rows with alternating colors
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2)):
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = light_fill
        
        elif template_name == "minimal":
            # Clean, minimal style
            thin_border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )
            
            header_font = Font(name="Segoe UI", size=11, bold=True)
            for cell in worksheet[1]:
                cell.font = header_font
                cell.border = thin_border
            
            # Add light borders to all cells
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border
    
    def _add_metadata_sheet(self, writer: pd.ExcelWriter, query: str, sql_query: str, data: List[Dict[str, Any]]):
        """Add a metadata sheet with query information"""
        metadata = [
            ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Natural Language Query", query],
            ["SQL Query", sql_query],
            ["Number of Records", len(data)],
            ["Columns", ", ".join(data[0].keys()) if data else "None"]
        ]
        
        metadata_df = pd.DataFrame(metadata, columns=["Metadata", "Value"])
        metadata_df.to_excel(writer, sheet_name="Metadata", index=False)
        
        # Format metadata sheet
        worksheet = writer.sheets["Metadata"]
        for i, _ in enumerate(metadata):
            # Make SQL query cell taller to accommodate multi-line text
            if i == 2:  # SQL Query row
                worksheet.row_dimensions[i+2].height = 60  # Adjust height
                
        # Auto-adjust column width
        for i, col in enumerate(["A", "B"]):
            max_length = max(
                len(str(metadata_df.iloc[j, i])) for j in range(len(metadata_df))
            )
            adjusted_width = min(max_length + 2, 100)  # Cap width at 100
            worksheet.column_dimensions[col].width = adjusted_width 